from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from config import Config
from models import db, Usuario, Cliente, Venta, VentaItem, Categoria, Producto, Variacion, InvoiceTemplate
from datetime import datetime
from pdf_service import generar_pdf_html, generar_pdf_boleta
from sunat_service import SUNATService
from mipse_service import MiPSEService
from scheduler_service import SchedulerService
import requests
import os
import atexit
import base64

app = Flask(__name__)
app.config.from_object(Config)

# Inicializar extensiones
db.init_app(app)

# LOG DE DEPURACI√ìN DE BASE DE DATOS (Muestra solo el Host por seguridad)
with app.app_context():
    db_host = app.config.get('SQLALCHEMY_DATABASE_URI').split('@')[-1].split(':')[0]
    print(f"üì° INTENTANDO CONEXI√ìN A DB EN: {db_host}")
    db.create_all()
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Globales para el scheduler y protecci√≥n de concurrencia
scheduler = None
_lock_file_handle = None  # Mantener vivo para que fcntl no libere el lock

def iniciar_scheduler():
    """Inicia el servicio de tareas programadas"""
    global scheduler
    if scheduler is None:
        scheduler = SchedulerService(app, db, Venta, MiPSEService)
        scheduler.iniciar()

def detener_scheduler():
    """Detiene el scheduler al cerrar la aplicaci√≥n"""
    global scheduler
    if scheduler:
        scheduler.detener()

# Registrar detenci√≥n al salir
atexit.register(detener_scheduler)

# Inicializar scheduler solo en el proceso principal o el primer worker
# En producci√≥n (Gunicorn), esto asegura que solo una instancia corra
def init_scheduler_production():
    def _do_init():
        global _lock_file_handle
        try:
            # Solo en Linux/Unix (Producci√≥n)
            import fcntl
            import time
            # Darle unos segundos adicionales para que el worker se estabilice
            time.sleep(5) 
            
            lock_file = os.path.join(os.getcwd(), '.scheduler.lock')
            _lock_file_handle = open(lock_file, 'wb')
            try:
                # Intentar bloquear el archivo
                fcntl.flock(_lock_file_handle, fcntl.LOCK_EX | fcntl.LOCK_NB)
                iniciar_scheduler()
                print("üöÄ Scheduler iniciado en este worker (background)")
            except OSError:
                # El lock ya lo tiene otro worker (u otro proceso), no hacemos nada
                _lock_file_handle.close()
                _lock_file_handle = None
        except ImportError:
            # En Windows (Desarrollo)
            if os.environ.get('WERKZEUG_RUN_MAIN') == 'true' or os.environ.get('FLASK_DEBUG') == '1':
                iniciar_scheduler()
        except Exception as e:
            print(f"‚ö†Ô∏è Error al inicializar scheduler: {e}")

    # Ejecutar en un hilo separado para no bloquear el arranque de Gunicorn
    import threading
    thread = threading.Thread(target=_do_init, daemon=True)
    thread.start()

@app.route('/health')
def health_check():
    """Ruta simple para health checks de Easypanel/Docker"""
    return jsonify({'status': 'healthy', 'timestamp': datetime.now().isoformat()}), 200

@login_manager.user_loader
def load_user(user_id):
    return Usuario.query.get(int(user_id))

# ==================== RUTAS DE AUTENTICACI√ìN ====================

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        identifier = request.form.get('login_identifier')
        password = request.form.get('password')
        
        # Buscar por email o username
        usuario = Usuario.query.filter(
            (Usuario.email == identifier) | (Usuario.username == identifier)
        ).first()
        
        if usuario and usuario.check_password(password):
            if not usuario.activo:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return jsonify({'success': False, 'message': 'Tu cuenta est√° desactivada'})
                flash('Tu cuenta est√° desactivada', 'danger')
                return render_template('login.html')
            
            login_user(usuario)
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': True, 'redirect': url_for('dashboard')})
            return redirect(url_for('dashboard'))
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'Credenciales incorrectas'})
            flash('Credenciales incorrectas', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        
        # Validar dominio
        if not email.endswith('@izistoreperu.com'):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'Solo se permiten correos @izistoreperu.com'})
            flash('Solo se permiten usuarios con correo @izistoreperu.com', 'danger')
            return render_template('login.html')
        
        # Verificar si email ya existe
        if Usuario.query.filter_by(email=email).first():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'El correo ya est√° registrado'})
            flash('El correo ya est√° registrado', 'danger')
            return render_template('login.html')

        # Verificar si username ya existe
        if Usuario.query.filter_by(username=username).first():
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return jsonify({'success': False, 'message': 'El nombre de usuario ya est√° en uso'})
            flash('El nombre de usuario ya est√° en uso', 'danger')
            return render_template('login.html')
        
        # Crear usuario
        usuario = Usuario(nombre=nombre, username=username, email=email, es_admin=True)
        usuario.set_password(password)
        
        db.session.add(usuario)
        db.session.commit()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': True, 'message': 'Registro exitoso. Ahora puedes iniciar sesi√≥n.'})
            
        flash('Usuario registrado exitosamente', 'success')
        return redirect(url_for('login'))
    
    return render_template('login.html')

# ==================== DASHBOARD ====================

@app.route('/dashboard')
@login_required
def dashboard():
    from sqlalchemy import func
    from datetime import date
    
    # Estad√≠sticas b√°sicas
    total_ventas = Venta.query.count()
    ventas_pendientes = Venta.query.filter_by(estado='PENDIENTE').count()
    ventas_enviadas = Venta.query.filter_by(estado='ENVIADO').count()
    
    # C√°lculo para Sem√°foro RUS (Mes actual)
    today = date.today()
    first_day = today.replace(day=1)
    
    total_mes = db.session.query(func.sum(Venta.total)).filter(
        Venta.fecha_emision >= first_day
    ).scalar() or 0
    total_mes = float(total_mes)
    
    # L√≠mites RUS
    limite_cat1 = 5000.00
    limite_cat2 = 8000.00
    
    # √öltimas ventas
    ultimas_ventas = Venta.query.order_by(Venta.fecha_emision.desc()).limit(10).all()
    
    return render_template('dashboard.html', 
                         total_ventas=total_ventas,
                         ventas_pendientes=ventas_pendientes,
                         ventas_enviadas=ventas_enviadas,
                         ultimas_ventas=ultimas_ventas,
                         total_mes=total_mes,
                         limite_cat1=limite_cat1,
                         limite_cat2=limite_cat2,
                         now=datetime.now())

def get_customer_data(tipo, numero, nombre_fallback=None, timeout=4):
    """
    Busca un cliente localmente o en el API externo (ApisPeru).
    Retorna (cliente, encontrado_en_db_local)
    """
    try:
        # 1. Buscar primero en la base de datos local
        cliente = Cliente.query.filter_by(numero_documento=numero).first()
        if cliente:
            print(f" [CUSTOMER-HELPER] Cliente {numero} encontrado localmente")
            return cliente, True

        # 2. Si no existe y es DNI, intentar con ApisPeru
        # Solo DNI de 8 d√≠gitos para evitar errores con otros documentos
        if tipo == 'DNI' and len(str(numero)) == 8:
            print(f" [CUSTOMER-HELPER] Buscando {numero} en ApisPeru (timeout {timeout}s)...")
            try:
                token = app.config.get('APISPERU_TOKEN')
                dni_url = app.config.get('APISPERU_DNI_URL')
                
                if token and dni_url:
                    url = f"{dni_url}/{numero}?token={token}"
                    response = requests.get(url, timeout=timeout)
                    
                    if response.status_code == 200:
                        data = response.json()
                        if 'dni' in data:
                            print(f" [CUSTOMER-HELPER] Datos encontrados en API para {numero}")
                            nuevo_cliente = Cliente(
                                tipo_documento='DNI',
                                numero_documento=data['dni'],
                                nombres=data['nombres'],
                                apellido_paterno=data['apellidoPaterno'],
                                apellido_materno=data['apellidoMaterno']
                            )
                            db.session.add(nuevo_cliente)
                            db.session.flush() # Flush para tener ID sin commit final
                            return nuevo_cliente, False
                        else:
                            print(f" [CUSTOMER-HELPER] API no retorn√≥ 'dni' para {numero}: {data}")
                    else:
                        print(f" [CUSTOMER-HELPER] API retorn√≥ status {response.status_code} para {numero}")
            except Exception as api_err:
                print(f" [CUSTOMER-HELPER] ‚ö†Ô∏è Error llamando a API DNI: {str(api_err)}")

        # 3. Fallback: Crear manual si tenemos nombre_fallback
        if nombre_fallback:
            print(f" [CUSTOMER-HELPER] Creando cliente {numero} con nombre fallback: {nombre_fallback}")
            nuevo_cliente = Cliente(
                tipo_documento=tipo,
                numero_documento=numero,
                nombres=nombre_fallback,
                apellido_paterno="",
                apellido_materno=""
            )
            db.session.add(nuevo_cliente)
            db.session.flush()
            return nuevo_cliente, False

        return None, False

    except Exception as e:
        print(f" [CUSTOMER-HELPER] ‚ùå Error cr√≠tico: {str(e)}")
        import traceback
        traceback.print_exc()
        return None, False

# ==================== API CONSULTA DNI/RUC ====================

@app.route('/api/buscar-cliente/<tipo>/<numero>')
@login_required
def buscar_cliente(tipo, numero):
    print(f" [BUSCAR-DEBUG] {datetime.now()} -> Iniciando busqueda para {tipo}: {numero}")
    try:
        cliente, local = get_customer_data(tipo, numero)
        
        if cliente:
            if not local:
                db.session.commit() # Si lo trajo de la API, guardamos
                
            return jsonify({
                'success': True,
                'existe': local,
                'cliente': {
                    'id': cliente.id,
                    'nombre_completo': cliente.nombre_completo,
                    'numero_documento': cliente.numero_documento,
                    'tipo_documento': cliente.tipo_documento,
                    'direccion': cliente.direccion or ''
                }
            })
        
        # Si no se encontr√≥ ni en local ni en API (y no se pas√≥ fallback)
        return jsonify({
            'success': False, 
            'allow_manual': True, 
            'message': f'{tipo} no encontrado o servicio no disponible. Ingrese datos manualmente.'
        }), 404
                
    except Exception as e:
        print(f" [BUSCAR-DEBUG] {datetime.now()} -> ‚ùå Error critico en ruta: {str(e)}")
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/api/clientes', methods=['POST'])
@login_required
def api_registrar_cliente():
    try:
        data = request.json
        tipo = data.get('tipo_documento')
        numero = data.get('numero_documento')
        nombre = data.get('nombre')
        direccion = data.get('direccion', '')

        if not tipo or not numero or not nombre:
            return jsonify({'success': False, 'message': 'Faltan datos requeridos'}), 400

        # Verificar si ya existe (por si acaso)
        existente = Cliente.query.filter_by(numero_documento=numero).first()
        if existente:
            return jsonify({'success': True, 'id': existente.id, 'nombre': existente.nombre_completo})

        # Para entrada manual usamos nombres para todo el campo
        nuevo = Cliente(
            tipo_documento=tipo,
            numero_documento=numero,
            nombres=nombre,
            direccion=direccion
        )
        db.session.add(nuevo)
        db.session.commit()

        return jsonify({
            'success': True,
            'id': nuevo.id,
            'nombre': nuevo.nombre_completo
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# ==================== VENTAS ====================

@app.route('/ventas')
@login_required
def ventas_list():
    ventas = Venta.query.order_by(Venta.fecha_emision.desc()).all()
    return render_template('ventas_list.html', ventas=ventas)

@app.route('/nueva-venta', methods=['GET', 'POST'])
@login_required
def nueva_venta():
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            cliente_id = request.form.get('cliente_id')
            numero_orden = request.form.get('numero_orden')
            items_data = request.form.getlist('items[]')
            
            # Obtener la serie del config
            serie = app.config['SERIE_BOLETA']

            # Calcular correlativo POR SERIE (buscar el m√°ximo correlativo de esta serie espec√≠fica)
            # SUNAT requiere 8 d√≠gitos para el correlativo
            max_correlativo = db.session.query(db.func.max(db.cast(Venta.correlativo, db.Integer)))\
                .filter(Venta.serie == serie)\
                .scalar()

            correlativo = 1 if not max_correlativo else max_correlativo + 1
            correlativo_str = str(correlativo).zfill(8)  # 8 d√≠gitos seg√∫n SUNAT

            numero_completo = f"{serie}-{correlativo_str}"
            
            # Crear venta
            venta = Venta(
                numero_orden=numero_orden if numero_orden else None,
                serie=serie,
                correlativo=correlativo_str,
                numero_completo=numero_completo,
                cliente_id=cliente_id,
                vendedor_id=current_user.id,
                subtotal=0,
                total=0,
                estado='PENDIENTE'
            )
            
            db.session.add(venta)
            db.session.flush()
            
            # Agregar items
            total = 0
            items = request.form.getlist('producto_nombre[]')
            skus = request.form.getlist('producto_sku[]')
            cantidades = request.form.getlist('cantidad[]')
            precios = request.form.getlist('precio_unitario[]')
            
            for i in range(len(items)):
                cantidad = float(cantidades[i])
                precio = float(precios[i])
                sku = skus[i] if i < len(skus) else ''
                subtotal = cantidad * precio
                total += subtotal
                
                item = VentaItem(
                    venta_id=venta.id,
                    producto_nombre=items[i],
                    producto_sku=sku,
                    cantidad=cantidad,
                    precio_unitario=precio,
                    subtotal=subtotal
                )
                db.session.add(item)
            
            venta.subtotal = total
            venta.total = total
            
            db.session.commit()
            
            # --- ENV√çO INMEDIATO A SUNAT (Requerido por usuario) ---
            print(f" [SALE-FLOW] Venta {numero_completo} creada. Iniciando env√≠o a SUNAT...")
            try:
                service = MiPSEService()
                resultado = service.procesar_venta(venta)
                
                if resultado['success']:
                    venta.estado = 'ENVIADO'
                    venta.fecha_envio_sunat = datetime.now()
                    venta.mensaje_sunat = resultado.get('message')
                    venta.hash_cpe = resultado.get('hash')
                    venta.external_id = resultado.get('external_id')
                    
                    # Guardar archivos XML y CDR
                    print(f" [SALE-FLOW] Guardando archivos para venta {venta.id}...")
                    guardar_archivos_mipse(venta, resultado)
                    
                    db.session.commit()
                    print(f" [SALE-FLOW] ‚úÖ Venta {numero_completo} procesada y guardada.")
                    flash(f'Venta {numero_completo} generada y enviada a SUNAT exitosamente.', 'success')
                else:
                    print(f" [SALE-FLOW] ‚ùå Error en MiPSE: {resultado}")
                    flash(f'Venta {numero_completo} creada localmente, pero hubo un error con SUNAT: {resultado.get("message")}', 'warning')
            except Exception as sunat_err:
                print(f" [SALE-FLOW] ‚ùå Error en env√≠o directo SUNAT: {str(sunat_err)}")
                flash(f'Venta {numero_completo} creada, pero el env√≠o a SUNAT fall√≥. Verifique en el listado de ventas.', 'info')

            return redirect(url_for('ventas_list'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear la venta: {str(e)}', 'danger')
    
    return render_template('nueva_venta.html')

# ==================== VER DETALLES DE VENTA ====================

@app.route('/venta/<int:venta_id>')
@login_required
def ver_venta(venta_id):
    venta = Venta.query.get_or_404(venta_id)
    return render_template('detalle_venta.html', venta=venta)

# ==================== GENERAR Y DESCARGAR PDF ====================

@app.route('/venta/<int:venta_id>/pdf')
@login_required
def descargar_pdf(venta_id):
    venta = Venta.query.get_or_404(venta_id)

    # Crear carpeta de comprobantes si no existe
    os.makedirs(app.config['COMPROBANTES_PATH'], exist_ok=True)

    # Nombre del archivo seg√∫n formato solicitado: Numero de Orden + Numero de Serie - Numero de Comprobante
    # Ejemplo: 3224368590_B001-00000009.pdf
    filename = f"{venta.numero_orden}_{venta.serie}-{venta.correlativo}.pdf"


    pdf_path = os.path.join(app.config['COMPROBANTES_PATH'], filename)

    # Generar PDF (Usando motor estable legacy por ahora seg√∫n solicitud usuario)
    if generar_pdf_boleta(venta, pdf_path):
        # Guardar la ruta en la base de datos
        venta.pdf_path = pdf_path
        db.session.commit()

        # Enviar el archivo (mostrar en navegador para imprimir)
        return send_file(pdf_path, as_attachment=False)
    else:
        flash('Error al generar el PDF', 'danger')
        return redirect(url_for('ventas_list'))

def guardar_archivos_mipse(venta, resultado):
    """Helper para decodificar y guardar archivos XML y CDR de MiPSE"""
    try:
        folder = app.config.get('COMPROBANTES_PATH', 'comprobantes')
        if not os.path.exists(folder):
            os.makedirs(folder, exist_ok=True)
            print(f" [SAVE-FILES] Carpeta creada: {folder}")
        
        nombre_base = resultado.get('nombre_archivo')
        if not nombre_base:
            tipo_doc = "03" if not (venta.serie and venta.serie.startswith('F')) else "01"
            correlativo = str(venta.correlativo).zfill(8)
            nombre_base = f"{app.config['EMPRESA_RUC']}-{tipo_doc}-{venta.serie}-{correlativo}"
            print(f" [SAVE-FILES] Usando nombre fallback: {nombre_base}")
        
        # Guardar XML Firmado
        xml_base64 = resultado.get('xml_firmado')
        if xml_base64:
            try:
                xml_path = os.path.join(folder, f"{nombre_base}.xml")
                with open(xml_path, 'wb') as f:
                    f.write(base64.b64decode(xml_base64))
                venta.xml_path = xml_path
                print(f" [SAVE-FILES] XML guardado: {xml_path}")
            except Exception as xml_err:
                print(f" [SAVE-FILES] ‚ùå Error escribiendo XML: {str(xml_err)}")
        else:
            print(f" [SAVE-FILES] ‚ö† No hay XML firmado en el resultado")
            
        # Guardar CDR
        cdr_base64 = resultado.get('cdr')
        if cdr_base64:
            try:
                # MiPSE devuelve el XML del CDR directamente, no un ZIP
                cdr_path = os.path.join(folder, f"R-{nombre_base}.xml")
                with open(cdr_path, 'wb') as f:
                    f.write(base64.b64decode(cdr_base64))
                venta.cdr_path = cdr_path
                print(f" [SAVE-FILES] CDR guardado: {cdr_path}")
            except Exception as cdr_err:
                print(f" [SAVE-FILES] ‚ùå Error escribiendo CDR: {str(cdr_err)}")
        else:
            print(f" [SAVE-FILES] ‚ö† No hay CDR en el resultado")
            
        return True
    except Exception as e:
        print(f" [SAVE-FILES] ‚ùå Error critico guardando archivos: {str(e)}")
        import traceback
        traceback.print_exc()
        return False

def recuperar_documentos_mipse(venta):
    """Intenta recuperar el CDR y XML desde MiPSE si no est√°n localmente"""
    try:
        service = MiPSEService()
        tipo_doc = "03" if not (venta.serie and venta.serie.startswith('F')) else "01"
        correlativo = str(venta.correlativo).zfill(8)
        nombre_archivo = f"{app.config['EMPRESA_RUC']}-{tipo_doc}-{venta.serie}-{correlativo}"
        
        print(f" [RECOVERY] Consultando estado en MiPSE para {nombre_archivo}...")
        resultado = service.consultar_estado(nombre_archivo)
        
        if resultado.get('success'):
            # Preparar datos para el helper de guardado
            data_files = {
                'nombre_archivo': nombre_archivo,
                'cdr': resultado.get('cdr'),
                'xml_firmado': resultado.get('xml_firmado') or resultado.get('data', {}).get('xml')
            }
            return guardar_archivos_mipse(venta, data_files)
        return False
    except Exception as e:
        print(f" [RECOVERY] ‚ùå Error recuperando documentos: {str(e)}")
        return False

# ==================== ENV√çO A SUNAT ====================

@app.route('/venta/<int:venta_id>/enviar-sunat', methods=['POST'])
@login_required
def enviar_sunat(venta_id):
    """Env√≠a la boleta a SUNAT"""
    venta = Venta.query.get_or_404(venta_id)

    # Verificar que no haya sido enviada antes
    if venta.estado == 'ENVIADO':
        flash('Esta venta ya fue enviada a SUNAT', 'warning')
        return redirect(url_for('ver_venta', venta_id=venta_id))

    try:
        # Inicializar servicio MiPSE (Proveedor autorizado)
        service = MiPSEService()

        # Procesar venta (generar XML via SUNATService localmente, firmar y enviar via MiPSE)
        resultado = service.procesar_venta(venta)

        if resultado['success']:
            # Actualizar estado de la venta
            venta.estado = 'ENVIADO'
            venta.fecha_envio_sunat = datetime.now()
            venta.mensaje_sunat = resultado.get('message')
            venta.hash_cpe = resultado.get('hash')
            venta.external_id = resultado.get('external_id')

            # Guardar archivos XML y CDR locales para descarga
            print(f" [MANUAL-SEND] Guardando archivos para venta {venta.id}...")
            guardar_archivos_mipse(venta, resultado)

            db.session.commit()
            print(f" [MANUAL-SEND] ‚úÖ Venta {venta.numero_completo} enviada e informaci√≥n guardada.")

            flash(f'Comprobante enviado exitosamente a SUNAT', 'success')
        else:
            flash(f"Error al enviar a SUNAT: {resultado.get('message', 'Error desconocido')}", 'danger')

    except Exception as e:
        flash(f'Error al procesar el env√≠o: {str(e)}', 'danger')

    return redirect(url_for('ver_venta', venta_id=venta_id))

@app.route('/venta/<int:venta_id>/xml')
@login_required
def descargar_xml(venta_id):
    """Descarga el XML de la boleta"""
    venta = Venta.query.get_or_404(venta_id)

    if not venta.xml_path or not os.path.exists(venta.xml_path):
        # Intentar recuperar si el estado es ENVIADO
        if venta.estado in ['ENVIADO', 'ACEPTADO']:
            if recuperar_documentos_mipse(venta):
                db.session.commit()
            else:
                flash('El XML no se pudo recuperar de MiPSE', 'warning')
                return redirect(url_for('ver_venta', venta_id=venta_id))
        else:
            flash('El XML no est√° disponible (la venta no ha sido enviada)', 'warning')
            return redirect(url_for('ver_venta', venta_id=venta_id))

    return send_file(venta.xml_path, as_attachment=True)

@app.route('/venta/<int:venta_id>/cdr')
@login_required
def descargar_cdr(venta_id):
    """Descarga el CDR (Constancia de Recepci√≥n) de SUNAT"""
    venta = Venta.query.get_or_404(venta_id)

    if not venta.cdr_path or not os.path.exists(venta.cdr_path):
        # Intentar recuperar si el estado es ENVIADO
        if venta.estado in ['ENVIADO', 'ACEPTADO']:
            if recuperar_documentos_mipse(venta):
                db.session.commit()
            else:
                flash('El CDR no se pudo recuperar de MiPSE', 'warning')
                return redirect(url_for('ver_venta', venta_id=venta_id))
        else:
            flash('El CDR no est√° disponible (la venta no ha sido enviada)', 'warning')
            return redirect(url_for('ver_venta', venta_id=venta_id))

    return send_file(venta.cdr_path, as_attachment=True)

# ==================== ELIMINACI√ìN DE VENTAS ====================

@app.route('/venta/<int:venta_id>/eliminar', methods=['DELETE'])
@login_required
def eliminar_venta(venta_id):
    """Elimina una venta individual"""
    try:
        venta = Venta.query.get_or_404(venta_id)

        # No permitir eliminar ventas enviadas a SUNAT
        if venta.estado == 'ENVIADO':
            return jsonify({
                'success': False,
                'message': 'No se puede eliminar una venta enviada a SUNAT'
            }), 400

        numero_completo = venta.numero_completo

        # Eliminar archivos asociados
        if venta.pdf_path and os.path.exists(venta.pdf_path):
            os.remove(venta.pdf_path)
        if venta.xml_path and os.path.exists(venta.xml_path):
            os.remove(venta.xml_path)

        # Eliminar venta (los items se eliminan en cascada)
        db.session.delete(venta)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Venta {numero_completo} eliminada exitosamente'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al eliminar la venta: {str(e)}'
        }), 500

@app.route('/ventas/eliminar-lote', methods=['DELETE'])
@login_required
def eliminar_lote():
    """Elimina m√∫ltiples ventas"""
    try:
        data = request.get_json()
        venta_ids = data.get('venta_ids', [])

        if not venta_ids:
            return jsonify({
                'success': False,
                'message': 'No se seleccionaron ventas'
            }), 400

        eliminadas = 0
        errores = []

        for venta_id in venta_ids:
            try:
                venta = Venta.query.get(venta_id)
                if not venta:
                    continue

                # No eliminar ventas enviadas a SUNAT
                if venta.estado == 'ENVIADO':
                    errores.append(f'{venta.numero_completo} (enviado a SUNAT)')
                    continue

                # Eliminar archivos asociados
                if venta.pdf_path and os.path.exists(venta.pdf_path):
                    os.remove(venta.pdf_path)
                if venta.xml_path and os.path.exists(venta.xml_path):
                    os.remove(venta.xml_path)

                db.session.delete(venta)
                eliminadas += 1

            except Exception as e:
                errores.append(f'Error en venta {venta_id}: {str(e)}')

        db.session.commit()

        mensaje = f'{eliminadas} venta(s) eliminada(s)'
        if errores:
            mensaje += f'. Errores: {", ".join(errores)}'

        return jsonify({
            'success': True,
            'eliminadas': eliminadas,
            'message': mensaje
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al eliminar ventas: {str(e)}'
        }), 500

# ==================== ENV√çO EN LOTE ====================

@app.route('/ventas/enviar-lote', methods=['POST'])
@login_required
def enviar_lote():
    """Env√≠a m√∫ltiples ventas a SUNAT"""
    try:
        data = request.get_json()
        venta_ids = data.get('venta_ids', [])

        if not venta_ids:
            return jsonify({
                'success': False,
                'message': 'No se seleccionaron ventas'
            }), 400

        service = MiPSEService()
        enviadas = 0
        errores = []

        for venta_id in venta_ids:
            try:
                venta = Venta.query.get(venta_id)
                if not venta:
                    continue

                # Solo enviar ventas pendientes
                if venta.estado != 'PENDIENTE':
                    continue

                # Procesar venta con MiPSE
                resultado = service.procesar_venta(venta)

                if resultado['success']:
                    venta.estado = 'ENVIADO'
                    venta.fecha_envio_sunat = datetime.utcnow()
                    venta.cdr_path = resultado.get('cdr_path')
                    venta.mensaje_sunat = resultado.get('message')
                    enviadas += 1
                else:
                    errores.append(f"{venta.numero_completo}: {resultado.get('message', 'Error desconocido')}")

            except Exception as e:
                errores.append(f'{venta.numero_completo}: {str(e)}')

        db.session.commit()

        mensaje = f'{enviadas} venta(s) enviada(s)'
        if errores:
            mensaje += f'. Errores: {", ".join(errores[:3])}'  # Mostrar solo 3 primeros errores

        return jsonify({
            'success': True,
            'enviadas': enviadas,
            'message': mensaje
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error al enviar ventas: {str(e)}'
        }), 500

# ==================== CONTROL DE SCHEDULER ====================

@app.route('/admin/scheduler/estado')
@login_required
def scheduler_estado():
    """Obtiene el estado del scheduler"""
    global scheduler
    if scheduler:
        estado = scheduler.obtener_estado()
        return jsonify(estado)
    return jsonify({'activo': False})

@app.route('/admin/scheduler/ejecutar-ahora', methods=['POST'])
@login_required
def scheduler_ejecutar_ahora():
    """Ejecuta el env√≠o autom√°tico inmediatamente (para pruebas)"""
    global scheduler
    if not scheduler:
        return jsonify({
            'success': False,
            'message': 'El scheduler no est√° activo'
        }), 400

    try:
        scheduler.ejecutar_ahora()
        return jsonify({
            'success': True,
            'message': 'Env√≠o autom√°tico ejecutado exitosamente'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error al ejecutar: {str(e)}'
        }), 500


@app.route('/api/get-categories')
@login_required
def get_categories():
    """Obtener √°rbol de categor√≠as de productos (desde DB local)"""
    try:
        # Obtener todas las categor√≠as
        all_categories = Categoria.query.order_by(Categoria.nombre).all()
        
        categories = []
        category_dict = {}

        # Primera pasada: crear todas las categor√≠as en el diccionario
        for cat in all_categories:
            cat_data = {
                'id': cat.id,
                'name': cat.nombre,
                'slug': cat.slug,
                'parent_id': cat.padre_id or 0,
                'product_count': cat.count,
                'children': []
            }
            category_dict[cat.id] = cat_data

            if cat.padre_id is None:
                categories.append(cat_data)

        # Segunda pasada: anidar subcategor√≠as
        for cat_id, cat_data in category_dict.items():
            parent_id = cat_data['parent_id']
            if parent_id != 0 and parent_id in category_dict:
                category_dict[parent_id]['children'].append(cat_data)

        return jsonify({
            'success': True,
            'categories': categories
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/get-products-by-category/<int:category_id>')
@login_required
def get_products_by_category(category_id):
    """Obtener productos de una categor√≠a espec√≠fica (desde DB local usando M2M)"""
    try:
        # Si category_id es 0, traer todos
        if category_id == 0:
            products_query = Producto.query.limit(100).all()
        else:
            # Filtrar usando la relaci√≥n muchos-a-muchos
            products_query = Producto.query.filter(
                Producto.categorias.any(Categoria.id == category_id)
            ).all()

        products = [{
            'id': p.id,
            'name': p.nombre,
            'sku': p.sku or 'N/A',
            'price': float(p.precio),
            'stock_status': p.stock_status,
            'imagen_url': p.imagen_url,
            'tipo': p.tipo
        } for p in products_query]

        return jsonify({
            'success': True,
            'products': products
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/search-products')
@login_required
def search_products():
    """Buscar productos por nombre o SKU localmente (incluye variaciones)"""
    q = request.args.get('q', '').strip()
    category_id = request.args.get('category_id', '0')
    
    if not q and (not category_id or category_id == '0'):
        return jsonify({'success': True, 'products': []})
    
    try:
        search_query = f"%{q}%"
        query = Producto.query
        
        if q:
            # Buscar en nombre del producto, SKU del producto principal O SKU de sus variaciones
            query = query.filter(
                (Producto.nombre.ilike(search_query)) | 
                (Producto.sku.ilike(search_query)) |
                (Producto.variaciones.any(Variacion.sku.ilike(search_query)))
            )
            
        # Si NO hay b√∫squeda por texto, filtramos por categor√≠a de forma estricta.
        # Si HAY b√∫squeda por texto, la categor√≠a es opcional (b√∫squeda global).
        if category_id and category_id != '0' and not q:
            query = query.filter(
                Producto.categorias.any(Categoria.id == int(category_id))
            )
            
        products_query = query.limit(50).all()

        products = [{
            'id': p.id,
            'name': p.nombre,
            'sku': p.sku or 'N/A',
            'price': float(p.precio),
            'stock_status': p.stock_status,
            'imagen_url': p.imagen_url,
            'tipo': p.tipo
        } for p in products_query]

        return jsonify({
            'success': True,
            'products': products
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/get-variations/<int:product_id>')
@login_required
def get_variations(product_id):
    """Obtener todas las variaciones de un producto"""
    try:
        from models import Variacion
        variations = Variacion.query.filter_by(producto_id=product_id).all()
        
        result = [{
            'id': v.id,
            'sku': v.sku,
            'price': float(v.precio),
            'stock_status': v.stock_status,
            'imagen_url': v.imagen_url,
            'atributos': v.atributos
        } for v in variations]
        
        return jsonify({
            'success': True,
            'variations': result
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/admin/ventas/bulk-upload', methods=['GET', 'POST'])
@login_required
def bulk_upload():
    """M√≥dulo de carga masiva desde Excel"""
    if request.method == 'GET':
        return render_template('bulk_upload.html')
    
    if 'file' not in request.files:
        flash('No se seleccion√≥ ning√∫n archivo', 'error')
        return redirect(request.url)
    
    file = request.files['file']
    if file.filename == '':
        flash('Archivo no v√°lido', 'error')
        return redirect(request.url)

    try:
        import pandas as pd
        import io
        
        # Leer Excel
        df = pd.read_excel(io.BytesIO(file.read()), engine='openpyxl')
        
        # Columnas: B(1)=SKU, E(4)=Order, L(11)=DNI, J(9)=Nombre, AJ(35)=Precio
        # Agrupar por Numero de Orden (Columna E)
        orders_dict = {}
        
        for index, row in df.iterrows():
            try:
                # Extraer datos usando iloc (√≠ndices 0-based)
                sku_raw = str(row.iloc[1]).strip() if not pd.isna(row.iloc[1]) else ""
                order_num = str(row.iloc[4]).strip() if not pd.isna(row.iloc[4]) else ""
                dni = str(row.iloc[11]).strip() if not pd.isna(row.iloc[11]) else "00000000"
                nombre = str(row.iloc[9]).strip() if not pd.isna(row.iloc[9]) else "CLIENTE VARIOS"
                precio = float(row.iloc[35]) if not pd.isna(row.iloc[35]) else 0.0
                desc_excel = str(row.iloc[40]).strip() if not pd.isna(row.iloc[40]) else ""
                costo_envio = float(row.iloc[37]) if not pd.isna(row.iloc[37]) else 0.0
                
                if not order_num: continue
                
                # Omitir si la orden ya existe en el sistema
                if order_num not in orders_dict:
                    existe = Venta.query.filter_by(numero_orden=order_num).first()
                    if existe:
                        # Marcamos como ignorada internamente para no procesar sus filas
                        orders_dict[order_num] = "EXISTE"
                        continue
                
                if orders_dict.get(order_num) == "EXISTE":
                    continue
                
                if order_num not in orders_dict:
                    orders_dict[order_num] = {
                        'order_num': order_num,
                        'dni': dni,
                        'nombre': nombre,
                        'order_items': [],
                        'total': 0,
                        'costo_envio': costo_envio,
                        'status': 'OK',
                        'errors': []
                    }
                
                # Buscar Match de SKU
                match_data = match_sku_intelligent(sku_raw)
                
                orders_dict[order_num]['order_items'].append({
                    'sku_excel': sku_raw,
                    'precio': precio,
                    'match': match_data,
                    'desc_excel': desc_excel
                })
                orders_dict[order_num]['total'] += precio
                
                if match_data['status'] == 'error':
                    orders_dict[order_num]['status'] = 'ERROR'
                    orders_dict[order_num]['errors'].append(f"SKU {sku_raw}: {match_data['msg']}")
                elif match_data['status'] == 'warning':
                    if orders_dict[order_num]['status'] == 'OK':
                        orders_dict[order_num]['status'] = 'WARNING'
                    orders_dict[order_num]['errors'].append(f"SKU {sku_raw}: {match_data['msg']}")
                    
            except Exception as e:
                print(f"Error procesando fila {index}: {e}")

        # Limpiar entradas marcadas como EXISTE y agregar costo de env√≠o al total
        final_orders = []
        for o in orders_dict.values():
            if isinstance(o, dict):
                # Sumar costo de env√≠o al total de la orden
                o['total'] += o.get('costo_envio', 0)
                final_orders.append(o)
        
        # Calcular Total del Mes actual para Monitor RUS
        from sqlalchemy import func
        from datetime import date
        today = date.today()
        first_day = today.replace(day=1)
        total_mes = db.session.query(func.sum(Venta.total)).filter(
            Venta.fecha_emision >= first_day
        ).scalar() or 0
        total_mes = float(total_mes)

        return render_template('bulk_upload_preview.html', 
                             orders=final_orders,
                             total_mes_actual=total_mes,
                             rus_cat1=5000.00,
                             rus_cat2=8000.00
                             )

    except Exception as e:
        flash(f'Error al procesar el Excel: {str(e)}', 'error')
        return redirect(request.url)

def match_sku_intelligent(sku_excel):
    """L√≥gica para emparejar SKU con BD (Exacto -> Parcial)"""
    if not sku_excel:
        return {'status': 'error', 'msg': 'SKU vac√≠o', 'id': None, 'name': 'N/A', 'type': 'simple'}
    
    # Normalizar: quitar espacios, convertir a may√∫sculas y limpiar artefactos de Excel (.0)
    sku_clean = str(sku_excel).strip()
    if sku_clean.endswith('.0'):
        sku_clean = sku_clean[:-2]
    
    # Normalizar guiones (en-dash, em-dash a guion normal)
    sku_clean = sku_clean.replace('‚Äì', '-').replace('‚Äî', '-')

    # 1. Match Exacto (Variaci√≥n) - Case Insensitive
    v = Variacion.query.filter(Variacion.sku.ilike(sku_clean)).first()
    if v:
        attr_desc = ", ".join([f"{k}: {v}" for k, v in v.atributos.items()])
        name = f"{v.producto.nombre} ({attr_desc})" if attr_desc else v.producto.nombre
        return {'status': 'ok', 'id': v.id, 'name': name, 'type': 'variable', 'sku': v.sku}
    
    # 2. Match Exacto (Producto) - Case Insensitive
    p = Producto.query.filter(Producto.sku.ilike(sku_clean)).first()
    if p:
        return {'status': 'ok', 'id': p.id, 'name': p.nombre, 'type': 'simple', 'sku': p.sku}

    # 3. Match por los primeros 7 d√≠gitos (si es compuesto o tiene sufijo)
    import re
    match_digits = re.search(r'(\d{7})', sku_clean)
    if match_digits:
        digits = match_digits.group(1)
        # Buscar variaciones que contengan esos d√≠gitos
        v_results = Variacion.query.filter(Variacion.sku.ilike(f"%{digits}%")).all()
        if len(v_results) == 1:
            v = v_results[0]
            attr_desc = ", ".join([f"{k}: {v}" for k, v in v.atributos.items()])
            name = f"{v.producto.nombre} ({attr_desc})" if attr_desc else v.producto.nombre
            return {'status': 'ok', 'id': v.id, 'name': name, 'type': 'variable', 'sku': v.sku}
        elif len(v_results) > 1:
            # Si hay m√∫ltiples, pero uno coincide exactamente con los primeros 7 d√≠gitos 
            # (Ej: 1007552 vs 1007552-Rojo), podr√≠amos ser m√°s espec√≠ficos.
            # Por ahora mantenemos la advertencia para que el usuario valide.
            return {'status': 'warning', 'msg': 'M√∫ltiples variaciones encontradas para la base', 'choices': len(v_results)}
            
        # Buscar productos que contengan esos d√≠gitos
        p_results = Producto.query.filter(Producto.sku.ilike(f"%{digits}%")).all()
        if len(p_results) == 1:
            p = p_results[0]
            return {'status': 'ok', 'id': p.id, 'name': p.nombre, 'type': 'simple', 'sku': p.sku}
        elif len(p_results) > 1:
            return {'status': 'warning', 'msg': 'M√∫ltiples productos encontrados para la base', 'choices': len(p_results)}

    return {'status': 'error', 'msg': 'No se encontr√≥ coincidencia', 'id': None, 'name': 'No encontrado', 'type': 'unknown'}

@app.route('/admin/ventas/bulk-process', methods=['POST'])
@login_required
def bulk_process():
    """Crear ventas reales desde el preview confirmado"""
    import json
    try:
        data = request.json
        orders = data.get('orders', [])
        
        results = {'success': 0, 'errors': 0}
        
        for o in orders:
            try:
                # 0. Verificaci√≥n de seguridad de √∫ltimo segundo (Duplicados)
                existe = Venta.query.filter_by(numero_orden=o['order_num']).first()
                if existe:
                    results['success'] += 1 # Contamos como procesado si ya existe
                    continue

                # 1. Buscar o Crear Cliente (Inteligente con API)
                # Usamos timeout de 3s para no colgar la carga masiva
                cliente, _ = get_customer_data('DNI', o['dni'], nombre_fallback=o['nombre'], timeout=3)
                
                if not cliente:
                    # Caso extremo si fall√≥ todo
                    results['errors'] += 1
                    continue
                
                # 2. Generar Correlativo (Boleta)
                serie = app.config.get('SERIE_BOLETA', 'B001')
                ultima_venta = Venta.query.filter_by(serie=serie).order_by(db.cast(Venta.correlativo, db.Integer).desc()).first()
                correlativo = 1
                if ultima_venta:
                    try:
                        correlativo = int(ultima_venta.correlativo) + 1
                    except:
                        correlativo = 1
                
                correlativo_str = str(correlativo).zfill(8)
                numero_completo = f"{serie}-{correlativo_str}"
                
                venta = Venta(
                    serie=serie,
                    correlativo=correlativo_str,
                    numero_completo=numero_completo,
                    numero_orden=o['order_num'],
                    cliente_id=cliente.id,
                    vendedor_id=current_user.id,
                    total=o['total'],
                    subtotal=o['total'],
                    estado='PENDIENTE'
                )
                db.session.add(venta)
                db.session.flush()
                
                # 3. Items
                for i in o['order_items']:
                    match = i['match']
                    item = VentaItem(
                        venta_id=venta.id,
                        producto_nombre=match['name'],
                        producto_sku=match['sku'],
                        cantidad=1,
                        precio_unitario=i['precio'],
                        subtotal=i['precio']
                    )
                    if match['type'] == 'variable':
                        item.variacion_id = match['id']
                    
                    db.session.add(item)
                
                db.session.commit()
                
                # 4. Enviar a SUNAT (MiPSE)
                try:
                    from mipse_service import MiPSEService
                    from app import guardar_archivos_mipse
                    
                    service = MiPSEService()
                    # Usar procesar_venta que maneja todo el flujo (XML + Firma + Env√≠o)
                    resultado = service.procesar_venta(venta)
                    
                    if resultado['success']:
                        venta.estado = 'ENVIADO'
                        venta.fecha_envio_sunat = datetime.now()
                        venta.mensaje_sunat = resultado.get('message')
                        venta.hash_cpe = resultado.get('hash')
                        venta.external_id = resultado.get('external_id')
                        
                        # Guardar archivos locales
                        try:
                            guardar_archivos_mipse(venta, resultado)
                        except Exception as file_err:
                            print(f"Error guardando archivos locales: {file_err}")
                            
                        db.session.commit()
                    else:
                        print(f"Error SUNAT para {numero_completo}: {resultado.get('error')}")
                        
                except Exception as sunat_err:
                    print(f"Error cr√≠tico enviando a SUNAT {numero_completo}: {sunat_err}")
                
                results['success'] += 1
                
            except Exception as order_err:
                db.session.rollback()
                print(f"Error creando orden {o['order_num']}: {order_err}")
                results['errors'] += 1
        
        return jsonify({'success': True, 'report': results})

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/admin/ventas/download-bulk', methods=['POST'])
@login_required
def download_bulk():
    """Descarga archivos masivos (PDF, XML, CDR) en un ZIP"""
    import zipfile
    import io
    import json
    
    try:
        venta_ids_json = request.form.get('venta_ids')
        tipo_archivo = request.form.get('tipo_archivo')
        
        if not venta_ids_json or not tipo_archivo:
            flash('Datos incompletos para la descarga', 'error')
            return redirect(url_for('ventas_list'))
            
        venta_ids = json.loads(venta_ids_json)
        ventas = Venta.query.filter(Venta.id.in_(venta_ids)).all()
        
        if not ventas:
            flash('No se encontraron ventas con los IDs proporcionados', 'error')
            return redirect(url_for('ventas_list'))

        # Crear ZIP en memoria
        memory_file = io.BytesIO()
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
            archivos_agregados = 0
            for v in ventas:
                filepath = None
                filename = None
                
                if tipo_archivo == 'pdf':
                    # Verificar si existe el PDF, si no generarlo
                    folder = app.config['COMPROBANTES_PATH']
                    os.makedirs(folder, exist_ok=True)
                    
                    filename = f"{v.numero_orden}_{v.serie}_{v.correlativo}.pdf"
                    expected_path = os.path.join(folder, filename)
                    
                    # Si no tiene ruta guardada O el archivo no existe f√≠sicamente
                    if not v.pdf_path or not os.path.exists(v.pdf_path if os.path.isabs(v.pdf_path) else os.path.join(app.root_path, v.pdf_path)):
                        print(f"Generando PDF faltante para venta {v.id}: {filename}")
                        if generar_pdf_boleta(v, expected_path):
                            v.pdf_path = expected_path
                            db.session.commit()
                            filepath = expected_path
                    else:
                        filepath = v.pdf_path

                    filename = f"{v.numero_orden}_{v.serie}-{v.correlativo}.pdf"
                elif tipo_archivo == 'xml':
                    filepath = v.xml_path
                    filename = f"{v.numero_orden}_{v.serie}-{v.correlativo}.xml"
                elif tipo_archivo == 'cdr':
                    filepath = v.cdr_path
                    filename = f"R-{v.numero_orden}_{v.serie}-{v.correlativo}.xml"
                
                if filepath:
                     # Si la ruta no es absoluta, la hacemos relativa a la raiz de la app
                     if not os.path.isabs(filepath):
                         filepath = os.path.join(app.root_path, filepath)

                     if os.path.exists(filepath):
                        # A√±adir al ZIP con el nombre limpio
                        zf.write(filepath, arcname=filename)
                        archivos_agregados += 1
        
        memory_file.seek(0)
        
        if archivos_agregados == 0:
            flash(f'No se encontraron archivos {tipo_archivo.upper()} para las ventas seleccionadas. Verifique que hayan sido enviadas a SUNAT.', 'warning')
            return redirect(url_for('ventas_list'))
            
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        zip_filename = f"ventas_{tipo_archivo}_{timestamp}.zip"
        
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=zip_filename
        )
            
    except Exception as e:
        print(f"Error en descarga masiva: {e}")
        flash(f'Error generando archivo ZIP: {str(e)}', 'error')
        return redirect(url_for('ventas_list'))

@app.route('/admin/ventas/download-errors', methods=['POST'])
@login_required
def download_bulk_errors():
    """Genera un archivo Excel con las √≥rdenes que tuvieron errores, ID√âNTICO al formato de carga"""
    import pandas as pd
    import io
    import json
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    
    try:
        data = request.json
        orders = data.get('orders', [])
        
        if not orders:
            return jsonify({'success': False, 'error': 'No hay datos para exportar'}), 400

        # Crear un libro de trabajo de OpenPyXL directamente para control total
        wb = Workbook()
        ws = wb.active
        ws.title = "Errores de Carga"
        
        # Definir encabezados y columnas clave (0-based indices)
        # B=1 (SKU), E=4 (Orden), J=9 (Nombre), L=11 (DNI), AJ=35 (Precio), AL=37 (Env√≠o), AO=40 (Desc), AP=41 (Error)
        col_map = {
            1: 'SKU',
            4: 'N¬∞ Orden',
            9: 'Nombre Cliente',
            11: 'DNI / RUC',
            35: 'Precio Unitario',
            37: 'Costo Env√≠o',
            40: 'Descripci√≥n',
            41: 'DETALLE DEL ERROR (Corregir y volver a subir)'
        }
        
        # Estilos
        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        key_col_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Amarillo suave
        error_col_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid") # Rojo suave
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Escribir encabezados en la fila 1
        for col_idx, title in col_map.items():
            cell = ws.cell(row=1, column=col_idx + 1, value=title)
            cell.font = bold_font
            cell.fill = header_fill
            cell.border = thin_border
            # Ajustar ancho
            ws.column_dimensions[cell.column_letter].width = 25

        current_row = 2
        
        for order in orders:
            # Una fila por cada item de la orden
            for item in order.get('order_items', []):
                
                # Extraer datos
                sku = item.get('sku_excel', '')
                orden = order.get('order_num', '')
                nombre = order.get('nombre', '')
                dni = order.get('dni', '')
                precio = item.get('precio', 0)
                desc = item.get('desc_excel', item.get('match', {}).get('name', ''))
                
                # Construir mensaje de error
                errors = order.get('errors', [])
                # Filtrar errores espec√≠ficos para este item si es posible, o mostrar todos los de la orden
                error_msg = ", ".join(errors) if errors else "Advertencia: Verificar datos"
                
                # Escribir en celdas espec√≠ficas
                # SKU (B)
                c_sku = ws.cell(row=current_row, column=2, value=sku)
                c_sku.fill = key_col_fill
                
                # Orden (E)
                c_ord = ws.cell(row=current_row, column=5, value=orden)
                c_ord.fill = key_col_fill
                
                # Nombre (J)
                c_nom = ws.cell(row=current_row, column=10, value=nombre)
                c_nom.fill = key_col_fill
                
                # DNI (L)
                c_dni = ws.cell(row=current_row, column=12, value=dni)
                c_dni.fill = key_col_fill
                
                # Precio (AJ) -> Columna 36
                c_pre = ws.cell(row=current_row, column=36, value=precio)
                c_pre.fill = key_col_fill
                
                # Costo Env√≠o (AL) -> Columna 38
                costo_envio = order.get('costo_envio', 0)
                c_env = ws.cell(row=current_row, column=38, value=costo_envio)
                c_env.fill = key_col_fill
                
                # Descripcion (AO) -> Columna 41
                c_desc = ws.cell(row=current_row, column=41, value=desc)
                c_desc.fill = key_col_fill
                
                # Error (AP) -> Columna 42
                c_err = ws.cell(row=current_row, column=42, value=error_msg)
                c_err.fill = error_col_fill
                c_err.font = Font(color="721C24", bold=True)
                
                current_row += 1

        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Convertir a base64
        import base64
        b64_data = base64.b64encode(output.read()).decode('utf-8')
        
        return jsonify({
            'success': True,
            'filename': f"errores_compatibles_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            'filedata': b64_data
        })

    except Exception as e:
        print(f"Error generando reporte de errores: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500



# ==================== DIS√ëADOR VISUAL DE PDF ====================

@app.route('/admin/diseno')
@login_required
def diseno_editor():
    # Obtener plantilla activa o crear una por defecto
    template = InvoiceTemplate.query.filter_by(es_activo=True).first()
    if not template:
        template = InvoiceTemplate.query.first()
        if not template:
            # Crear plantilla inicial por defecto (HTML b√°sico)
            template = InvoiceTemplate(
                nombre="A4 Est√°ndar",
                es_activo=True,
                html_content="""
                <div style="font-family: Arial; padding: 20px;">
                    <h1 style="text-align: center;">[[EMPRESA_NOMBRE]]</h1>
                    <p style="text-align: center;">RUC: [[EMPRESA_RUC]]</p>
                    <hr>
                    <div style="border: 1px solid black; padding: 10px; margin: 10px 0;">
                        <strong>BOLETA DE VENTA ELECTR√ìNICA</strong><br>
                        Nro: [[NRO_COMPROBANTE]]
                    </div>
                    <p>Cliente: [[CLIENTE_NOMBRE]]</p>
                    <p>Fecha: [[FECHA_EMISION]]</p>
                    <table style="width: 100%; border-collapse: collapse; margin-top: 20px;">
                        <thead>
                            <tr style="background: #eee;">
                                <th style="border: 1px solid black; padding: 5px;">DESCRIPCI√ìN</th>
                                <th style="border: 1px solid black; padding: 5px;">CANT.</th>
                                <th style="border: 1px solid black; padding: 5px;">P. UNIT</th>
                                <th style="border: 1px solid black; padding: 5px;">TOTAL</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td colspan="4" style="border: 1px solid black; padding: 20px; text-align: center;">
                                    [[DETALLE_PRODUCTOS]]
                                </td>
                            </tr>
                        </tbody>
                    </table>
                    <h3 style="text-align: right;">TOTAL: [[TOTAL]]</h3>
                </div>
                """,
                css_content="body { font-size: 12px; }"
            )
            db.session.add(template)
            db.session.commit()
    
    return render_template('diseno_editor.html', template=template)

@app.route('/api/diseno/guardar', methods=['POST'])
@login_required
def guardar_diseno():
    try:
        data = request.json
        html = data.get('html')
        css = data.get('css')
        
        template = InvoiceTemplate.query.filter_by(es_activo=True).first()
        if not template:
            template = InvoiceTemplate(nombre="Personalizada", es_activo=True)
            db.session.add(template)
        
        template.html_content = html
        template.css_content = css
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/diseno/preview')
@login_required
def diseno_preview():
    from models import Venta, VentaItem, Cliente, InvoiceTemplate
    from pdf_service import generar_pdf_html
    
    # 1. Obtener una venta real para el test (o crear un mock si no hay ninguna)
    venta = Venta.query.order_by(Venta.fecha_emision.desc()).first()
    
    if not venta:
        # Mock b√°sico para pruebas iniciales
        class MockVenta:
            def __init__(self):
                self.id = 0
                self.numero_completo = "B001-00000000"
                self.fecha_emision = datetime.now()
                self.serie = "B001"
                self.correlativo = "00000000"
                self.numero_orden = "123456789"
                self.total = 100.00
                self.hash_cpe = "ABCDEF123456"
                self.cliente = Cliente(nombre_completo="CLIENTE DE PRUEBA", numero_documento="77777777", tipo_documento="DNI", direccion="CALLE DE PRUEBA 123")
                self.items = [
                    VentaItem(producto_nombre="Producto de Prueba A", cantidad=2, precio_unitario=25, subtotal=50),
                    VentaItem(producto_nombre="Producto de Prueba B", cantidad=1, precio_unitario=50, subtotal=50)
                ]
                self.vendedor = None
        venta = MockVenta()

    # 2. Generar en un archivo temporal de preview
    preview_path = os.path.join(app.config['COMPROBANTES_PATH'], "preview_diseno.pdf")
    
    if generar_pdf_html(venta, preview_path):
        return send_file(preview_path, as_attachment=False, max_age=0)
    else:
        return "Error al generar la vista previa", 500


# Iniciar directorios
os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)
os.makedirs(Config.COMPROBANTES_PATH, exist_ok=True)

# Iniciar scheduler (con protecci√≥n de bloqueo en producci√≥n)
with app.app_context():
    init_scheduler_production()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
